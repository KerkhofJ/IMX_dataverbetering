from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def get_revision_template(out_file_path: str | Path):
    if isinstance(out_file_path, str):
        out_file_path = Path(out_file_path)

    columns = [
        "ObjectPath",
        "ObjectPuic",
        "IssueComment",
        "IssueCause",
        "AtributeOrElement",
        "Operation",
        "ValueOld",
        "ValueNew",
        "ProcessingStatus",
        "RevisionReasoning",
    ]
    first_row = [
        "Imx Object Path",
        "Puic of object for revision",
        "What is the issue",
        "What is the cause of the issue",
        "The attribute or element path",
        "Type of revision operation",
        "Old value that is being checked if it is still like this",
        "Revision Value",
        "Boolean If revision need to be processed",
        "Revision reasoning, why this value, or why not to revision?",
    ]
    example_row_1 = [
        "dummy.object.path",
        "a8cfb00e-bbb3-4a83-9783-4f94e013fa9d",
        "Unknown upgrade",
        "just a example",
        "@name",
        "UpdateAttribute",
        "Unknown",
        "SomeOtherValue",
        "True",
        "Will revision if old value still matches",
    ]

    example_row_2 = [
        "dumy.obkect.path",
        "a8cfb00e-bbb3-4a83-9783-4f94e013fa9d",
        "Unknown upgrade",
        "just a example",
        "RailConnectionInfo.@atMeasure",
        "UpdateAttribute",
        "0",
        "0.255",
        "False",
        "Will NOT revision ProcessingStatus = False",
    ]

    wb = Workbook()
    ws = wb.active
    if not ws:
        raise ValueError("Workbook not found")

    ws.title = "revisions"

    # set header
    for col_idx, value in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=value)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # first row should have column info
    for col_idx, value in enumerate(first_row, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    # first row should have column info
    for col_idx, value in enumerate(example_row_1, start=1):
        ws.cell(row=3, column=col_idx, value=value)

    # first row should have column info
    for col_idx, value in enumerate(example_row_2, start=1):
        ws.cell(row=4, column=col_idx, value=value)

    for col_idx in range(1, ws.max_column + 1):  # Loop through columns
        column = get_column_letter(col_idx)

        # Get the maximum length of the values in the column (including padding)
        max_length = max(
            (
                len(str(cell))
                for row in ws.iter_rows(
                    min_col=col_idx, max_col=col_idx, values_only=True
                )
                for cell in row
            ),
            default=0,  # In case there are empty columns, avoid errors
        )

        # Set the column width, adding padding for visibility
        ws.column_dimensions[column].width = max_length + 2

    operation_values = [
        "CreateAttribute",
        "UpdateAttribute",
        "DeleteAttribute",
        "DeleteObject",
        "AddElementUnder",
        "DeleteElement",
    ]
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(operation_values)}"',
        allow_blank=True,
        showDropDown=False,  # counterintuitive, false to show dropdown :/
    )
    ws.add_data_validation(dv)
    operation_col_letter = get_column_letter(columns.index("Operation") + 1)
    dv_range = f"{operation_col_letter}2:{operation_col_letter}1048576"
    dv.add(dv_range)

    operation_values = ["True", "False"]
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(operation_values)}"',
        allow_blank=True,
        showDropDown=False,  # counterintuitive, false to show dropdown :/
    )
    ws.add_data_validation(dv)
    operation_col_letter = get_column_letter(columns.index("ProcessingStatus") + 1)
    dv_range = f"{operation_col_letter}2:{operation_col_letter}1048576"
    dv.add(dv_range)

    wb.save(out_file_path)
