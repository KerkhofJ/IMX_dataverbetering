import os
import shutil

from imxInsights.utils.report_helpers import REVIEW_STYLES
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from settings import ISSUE_LIST_SHEET_NAME


def get_cell_background_color(cell: Cell | MergedCell) -> str | None:
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
        return cell.fill.fgColor.rgb[-6:]
    return None


def get_column_indices(ws: Worksheet, header_row: int) -> dict[str, int | None]:
    keys = ["@puic", "path", "status", "geometry_status"]
    return {
        key: next(
            (
                col
                for col in range(1, ws.max_column + 1)
                if ws[f"{get_column_letter(col)}{header_row}"].value == key
            ),
            None,
        )
        for key in keys
    }


def get_cell_context(
    ws: Worksheet, row_idx: int, columns: dict[str, int | None]
) -> dict[str, str | None]:
    def get(col_key: str) -> str | None:
        col = columns.get(col_key)
        val = ws.cell(row=row_idx, column=col).value if col else None
        return str(val) if val is not None else None

    return {
        "Puic": get("@puic"),
        "ObjectPath": get("path"),
        "ChangeStatus": get("status"),
        "GeometryStatus": get("geometry_status"),
    }


def build_comment_dict(
    context: dict[str, str | None],
    cell: Cell | MergedCell,
    header: str,
    sheet_name: str,
    comment_text: str,
) -> dict[str, str | int | None]:
    return {
        "Sheet": sheet_name,
        "Puic": context["Puic"],
        "Header": header,
        "Value": str(cell.value) if cell.value is not None else "",
        "Comment": comment_text,
        "CellAddress": cell.coordinate,
        "Color": get_cell_background_color(cell),
        "ObjectPath": context["ObjectPath"],
        "ChangeStatus": context["ChangeStatus"],
        "GeometryStatus": context["GeometryStatus"],
        "CommentSheetName": sheet_name,
        "CommentRow": cell.row,
        "CommentColumn": cell.column,
    }


def handle_header_comment(
    cell: Cell | MergedCell,
    ws: Worksheet,
    header_value: str,
    columns: dict[str, int | None],
    sheet_name: str,
    header_row: int,
) -> tuple[list[dict[str, str | int | None]], list[str]]:
    comments: list[dict[str, str | int | None]] = []
    inherited_cells: list[str] = []

    if get_cell_background_color(cell) == "FFFFFF":
        return comments, inherited_cells

    for row_idx in range(header_row + 1, ws.max_row + 1):
        if cell.column is None:
            continue
        target_cell = ws.cell(row=row_idx, column=cell.column)

        if target_cell.comment or not target_cell.value:
            continue
        color = get_cell_background_color(target_cell)
        if color and color != "000000":
            context = get_cell_context(ws, row_idx, columns)
            if cell.comment is not None:
                comments.append(
                    build_comment_dict(
                        context,
                        target_cell,
                        header_value,
                        sheet_name,
                        cell.comment.text,
                    )
                )
                inherited_cells.append(target_cell.coordinate)
    return comments, inherited_cells


def handle_data_comment(
    cell: Cell | MergedCell,
    ws: Worksheet,
    header_value: str,
    columns: dict[str, int | None],
    sheet_name: str,
) -> list[dict[str, str | int | None]]:
    if cell.row is None:
        return []
    context = get_cell_context(ws, cell.row, columns)
    return [
        build_comment_dict(
            context,
            cell,
            header_value,
            sheet_name,
            cell.comment.text if cell.comment else "",
        )
    ]


def write_comments_sheet(
    wb: Workbook,
    comments: list[dict[str, str | int | None]],
    overwrite: bool = False,
) -> None:
    if "Comments" in wb.sheetnames:
        if overwrite:
            del wb[ISSUE_LIST_SHEET_NAME]
        else:
            raise ValueError(
                f"Sheet '{ISSUE_LIST_SHEET_NAME}' already exists. Set overwrite=True to overwrite."
            )

    # Create the new sheet
    ws_comments = wb.create_sheet(ISSUE_LIST_SHEET_NAME)

    # Reorder the sheet to be after 'info' or first if 'info' not present
    info_index = next(
        (i for i, ws in enumerate(wb.worksheets) if ws.title == "info"), None
    )
    worksheets = wb.worksheets
    worksheets.remove(ws_comments)
    insert_at = info_index + 1 if info_index is not None else 0
    worksheets.insert(insert_at, ws_comments)

    ws_comments.append(
        [
            "ObjectPath",
            "Puic",
            "ChangeStatus",
            "GeometryStatus",
            "Link",
            "ImxPath",
            "Value",
            "Comment",
            "Color",
            "CommentSheetName",
            "CommentRow",
            "CommentColumn",
        ]
    )
    ws_comments.auto_filter.ref = f"A1:{get_column_letter(ws_comments.max_column)}1"

    color_to_status = {v: k for k, v in REVIEW_STYLES.items()}

    for comment in comments:
        link_text = color_to_status.get(str(comment["Color"]), "link")
        link = f'=HYPERLINK("#\'{comment["Sheet"]}\'!{comment["CellAddress"]}", "{link_text}")'
        color = str(comment["Color"]) if comment["Color"] is not None else None
        row = [
            comment["ObjectPath"],
            comment["Puic"],
            comment["ChangeStatus"],
            comment["GeometryStatus"],
            link,
            comment["Header"],
            comment["Value"],
            comment["Comment"],
            comment["Color"],
            comment["CommentSheetName"],
            comment["CommentRow"],
            comment["CommentColumn"],
        ]
        ws_comments.append(row)
        if color:
            ws_comments[f"E{ws_comments.max_row}"].fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid",
            )

    ws_comments.freeze_panes = "A2"

    for col in range(1, ws_comments.max_column + 1):
        col_cells = next(ws_comments.iter_cols(min_col=col, max_col=col))
        max_len = max(len(str(cell.value) or "") for cell in col_cells) + 4
        ws_comments.column_dimensions[get_column_letter(col)].width = max_len


def extract_comments_to_new_sheet(
    file_path: str,
    output_path: str | None = None,
    header_row: int = 1,
    add_to_wb: bool = False,
    overwrite: bool = False,
) -> None:
    if not output_path and not add_to_wb:
        raise ValueError("When adding to an existing workbook, provide an output path.")

    target_path = output_path if output_path and add_to_wb else file_path
    if output_path and add_to_wb:
        if os.path.exists(output_path) and not overwrite:
            raise FileExistsError(
                f"File '{output_path}' already exists. Set overwrite=True."
            )
        shutil.copyfile(file_path, output_path)

    wb = load_workbook(target_path, data_only=True)
    all_comments: list[dict[str, str | int | None]] = []
    inherited_cells: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        columns = get_column_indices(ws, header_row)

        for row in ws.iter_rows():
            for cell in row:
                if cell.column is None:
                    continue
                header_cell = ws[f"{get_column_letter(cell.column)}{header_row}"]
                header_value = header_cell.value if header_cell else None
                header_value_str = str(header_value) if header_value is not None else ""

                if cell.comment:
                    if cell.row == header_row and isinstance(header_value, str):
                        comments, inherited = handle_header_comment(
                            cell, ws, header_value, columns, sheet_name, header_row
                        )
                        all_comments.extend(comments)
                        inherited_cells.update(inherited)
                    elif cell.row != header_row:
                        if cell.row is not None:
                            all_comments.extend(
                                handle_data_comment(
                                    cell, ws, header_value_str, columns, sheet_name
                                )
                            )
                else:
                    color = get_cell_background_color(cell)
                    if (
                        cell.row is not None
                        and cell.row > header_row
                        and color in REVIEW_STYLES.values()
                        and cell.value
                        and cell.coordinate not in inherited_cells
                    ):
                        context = get_cell_context(ws, cell.row, columns)
                        all_comments.append(
                            build_comment_dict(
                                context, cell, header_value_str, sheet_name, ""
                            )
                        )

    if add_to_wb:
        if os.path.exists(file_path) and not overwrite:
            raise FileExistsError(
                f"File '{file_path}' already exists. Set overwrite=True."
            )
        write_comments_sheet(wb, all_comments, overwrite)
        wb.save(target_path)
        print(
            f"Comments sheet written to '{target_path}' (in-place={target_path == file_path})."
        )
    else:
        if output_path and os.path.exists(output_path) and not overwrite:
            raise FileExistsError(
                f"File '{output_path}' already exists. Set overwrite=True."
            )
        wb_new = Workbook()
        active_sheet = wb_new.active
        if isinstance(active_sheet, Worksheet):
            wb_new.remove(active_sheet)

        write_comments_sheet(wb_new, all_comments, overwrite)
        output_path = output_path or file_path.replace(".xlsx", "_comments.xlsx")
        wb_new.save(output_path)
        print(
            f"Comments extracted successfully to '{output_path}' in sheet '{ISSUE_LIST_SHEET_NAME}'."
        )
