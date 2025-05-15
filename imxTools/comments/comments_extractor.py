import os
import shutil
from pathlib import Path

from imxInsights.utils.report_helpers import REVIEW_STYLES
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from imxTools.comments.comment_entry import CommentEntry
from imxTools.comments.comments_enums import CommentColumns

ISSUE_LIST_SHEET_NAME = "comments"
CellType = Cell | MergedCell


def get_cell_background_color(cell: CellType) -> str | None:
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
        return cell.fill.fgColor.rgb[-6:]
    return None


def get_column_indices(
    ws: Worksheet, header_row: int, keys: list[str] | None = None
) -> dict[str, int | None]:
    if keys is None:
        keys = ["@puic", "path", "status", "geometry_status"]

    return {
        key: next(
            (
                col
                for col in range(1, ws.max_column + 1)
                if ws[f"{get_column_letter(col)}{header_row}"].value == key
            ),
            None,
        )
        for key in keys
    }


def get_cell_context(
    ws: Worksheet, row_idx: int, columns: dict[str, int | None]
) -> dict[str, str]:
    def get(col_key: str) -> str:
        col = columns.get(col_key)
        val = ws.cell(row=row_idx, column=col).value if col else None
        return str(val) if val is not None else ""

    return {
        CommentColumns.object_puic: get("@puic"),
        CommentColumns.object_path: get("path"),
        CommentColumns.change_status: get("status"),
        CommentColumns.geometry_status: get("geometry_status"),
    }


def build_comment_entry(
    cell: CellType,
    header: str,
    sheet_name: str,
    comment_text: str,
    comment_row: int | None,
    context: dict[str, str],
) -> CommentEntry:
    return CommentEntry(
        sheet_name=sheet_name,
        header_value=header,
        cell_value=str(cell.value) if cell.value is not None else "",
        comment=comment_text,
        cell_address=cell.coordinate,
        bg_color=get_cell_background_color(cell),
        row=comment_row
        if comment_row is not None
        else cell.row
        if cell.row is not None
        else 0,
        column=cell.column if cell.column is not None else 0,
        puic=context[CommentColumns.object_puic],
        path=context[CommentColumns.object_path],
        status=context[CommentColumns.change_status],
        geometry_status=context[CommentColumns.geometry_status],
    )


def handle_header_comment(
    cell: CellType,
    ws: Worksheet,
    header_value: str,
    columns: dict[str, int | None],
    sheet_name: str,
    header_row: int,
) -> tuple[list[CommentEntry], list[str]]:
    comments: list[CommentEntry] = []
    inherited_cells: list[str] = []

    if get_cell_background_color(cell) == "FFFFFF":
        return comments, inherited_cells

    for row_idx in range(header_row + 1, ws.max_row + 1):
        if cell.column:
            target_cell = ws.cell(row=row_idx, column=cell.column)
            if target_cell.comment or not target_cell.value:
                continue
            color = get_cell_background_color(target_cell)
            if color and color != "000000":
                context = get_cell_context(ws, row_idx, columns)
                if cell.comment:
                    entry = build_comment_entry(
                        target_cell,
                        header_value,
                        sheet_name,
                        cell.comment.text,
                        header_row,
                        context,
                    )
                    comments.append(entry)
                    inherited_cells.append(target_cell.coordinate)

    return comments, inherited_cells


def handle_data_comment(
    cell: CellType,
    ws: Worksheet,
    header_value: str,
    columns: dict[str, int | None],
    sheet_name: str,
) -> list[CommentEntry]:
    if cell.row:
        context = get_cell_context(ws, cell.row, columns)
        entry = build_comment_entry(
            cell,
            header_value,
            sheet_name,
            cell.comment.text if cell.comment else "",
            None,
            context,
        )
        return [entry]
    return []


def write_comments_sheet(
    wb: Workbook,
    comments: list[CommentEntry],
    overwrite: bool = False,
) -> None:
    if ISSUE_LIST_SHEET_NAME in wb.sheetnames:
        if overwrite:
            del wb[ISSUE_LIST_SHEET_NAME]
        else:
            raise ValueError(
                f"Sheet '{ISSUE_LIST_SHEET_NAME}' already exists. Set overwrite=True to overwrite."
            )

    ws_comments = wb.create_sheet(ISSUE_LIST_SHEET_NAME)
    worksheets = wb.worksheets
    worksheets.remove(ws_comments)
    insert_at = next(
        (i + 1 for i, ws in enumerate(worksheets) if ws.title == "info"), 0
    )
    worksheets.insert(insert_at, ws_comments)

    ws_comments.append(CommentColumns.names())
    ws_comments.auto_filter.ref = f"A1:{get_column_letter(ws_comments.max_column)}1"

    color_to_status = {v: k for k, v in REVIEW_STYLES.items()}
    for entry in comments:
        link_text = color_to_status.get(str(entry.bg_color), "link")
        link = (
            f'=HYPERLINK("#\'{entry.sheet_name}\'!{entry.cell_address}", "{link_text}")'
        )
        row = [
            entry.path,
            entry.puic,
            entry.status,
            entry.geometry_status,
            link,
            entry.header_value,
            entry.cell_value,
            entry.comment,
            entry.bg_color,
            entry.sheet_name,
            entry.row,
            entry.column,
        ]
        ws_comments.append(row)
        if entry.bg_color:
            ws_comments[f"E{ws_comments.max_row}"].fill = PatternFill(
                start_color=entry.bg_color, end_color=entry.bg_color, fill_type="solid"
            )

    ws_comments.freeze_panes = "A2"

    for col in range(1, ws_comments.max_column + 1):
        col_cells = next(ws_comments.iter_cols(min_col=col, max_col=col))
        max_len = max(len(str(cell.value) or "") for cell in col_cells) + 4
        ws_comments.column_dimensions[get_column_letter(col)].width = max_len


def extract_comments_to_new_sheet(
    file_path: str | Path,
    output_path: str | None = None,
    header_row: int = 1,
    add_to_wb: bool = False,
    overwrite: bool = False,
) -> None:
    if not output_path and not add_to_wb:
        raise ValueError("When adding to an existing workbook, provide an output path.")

    target_path = output_path if output_path and add_to_wb else file_path
    if output_path and add_to_wb:
        if os.path.exists(output_path) and not overwrite:
            raise FileExistsError(
                f"File '{output_path}' already exists. Set overwrite=True."
            )
        shutil.copyfile(file_path, output_path)

    wb = load_workbook(target_path, data_only=True)
    all_comments: list[CommentEntry] = []
    inherited_cells: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        columns = get_column_indices(ws, header_row)
        if columns is None:
            continue

        for row in ws.iter_rows():
            for cell in row:
                if cell.column is None or cell.row is None:
                    continue

                col_letter = get_column_letter(cell.column)
                header_cell = ws[f"{col_letter}{header_row}"]
                header_value = str(header_cell.value or "")

                if cell.comment:
                    if cell.row == header_row:
                        comments, inherited = handle_header_comment(
                            cell, ws, header_value, columns, sheet_name, header_row
                        )
                        all_comments.extend(comments)
                        inherited_cells.update(inherited)
                    else:
                        all_comments.extend(
                            handle_data_comment(
                                cell, ws, header_value, columns, sheet_name
                            )
                        )
                else:
                    color = get_cell_background_color(cell)
                    if (
                        cell.row > header_row
                        and color in REVIEW_STYLES.values()
                        and cell.value
                        and cell.coordinate not in inherited_cells
                    ):
                        context = get_cell_context(ws, cell.row, columns)
                        entry = build_comment_entry(
                            cell, header_value, sheet_name, "", None, context
                        )
                        all_comments.append(entry)

    if add_to_wb:
        if os.path.exists(file_path) and not overwrite:
            raise FileExistsError(
                f"File '{file_path}' already exists. Set overwrite=True."
            )
        write_comments_sheet(wb, all_comments, overwrite)
        wb.save(target_path)
        print(
            f"Comments sheet written to '{target_path}' (in-place={target_path == file_path})."
        )
    else:
        output_path = output_path or str(file_path).replace(".xlsx", "_comments.xlsx")
        wb_new = Workbook()
        if wb_new.active:
            wb_new.remove(wb_new.active)
        write_comments_sheet(wb_new, all_comments, overwrite)
        wb_new.save(output_path)
        print(
            f"Comments extracted successfully to '{output_path}' in sheet '{ISSUE_LIST_SHEET_NAME}'."
        )
