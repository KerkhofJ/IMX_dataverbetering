from copy import copy
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from imxTools.settings import ISSUE_LIST_SHEET_NAME


def copy_full_sheet(source_ws: Worksheet, target_ws: Worksheet):
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.column)

            if cell.data_type == "f":  # Cell has a formula
                target_cell.value = f"={cell.value}"
            else:
                target_cell.value = cell.value

            if cell.hyperlink:
                target_cell.hyperlink = copy(cell.hyperlink)

            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.fill = copy(cell.fill)
                target_cell.border = copy(cell.border)
                target_cell.alignment = copy(cell.alignment)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)

    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width

    for row_idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = dim.height

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))


def get_sheet_headers(ws, header_row: int = 1) -> dict:
    return {
        cell.value: idx
        for idx, cell in enumerate(
            next(ws.iter_rows(min_row=header_row, max_row=header_row)), start=1
        )
    }


def find_column_by_value(ws, target: str, header_row: int) -> int | None:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == target:
            return col
    return None


def find_row_by_value(ws, col: int, value: Any, start_row: int) -> int | None:
    for row in range(start_row, ws.max_row + 1):
        if str(ws.cell(row=row, column=col).value) == str(value):
            return row
    return None


def create_summary_sheet(wb, processed, skipped, not_found):
    summary_ws = wb.create_sheet("CommentPlacementSummary")
    summary_ws.append(
        ["Status", "Sheet", "ImxPath", "Puic", "Value", "Comment", "Reason"]
    )

    for row in processed:
        summary_ws.append(
            [
                "Placed",
                row.get("CommentSheetName", ""),
                row.get("ImxPath", ""),
                row.get("Puic", ""),
                row.get("Value", ""),
                row.get("Comment", ""),
                "",
            ]
        )
    for row in skipped:
        summary_ws.append(
            [
                "Skipped",
                row.get("CommentSheetName", ""),
                row.get("ImxPath", ""),
                row.get("Puic", ""),
                row.get("Value", ""),
                row.get("Comment", ""),
                row.get("Reason", ""),
            ]
        )
    for row in not_found:
        summary_ws.append(
            [
                "Failed",
                row.get("CommentSheetName", ""),
                row.get("ImxPath", ""),
                row.get("Puic", ""),
                row.get("Value", ""),
                row.get("Comment", ""),
                row.get("Reason", ""),
            ]
        )
    return summary_ws


def extract_display_text(formula: str) -> str:
    parts = formula.split(",")
    if len(parts) >= 2:
        return parts[1].strip().rstrip(')"').strip('"')
    return ""


def apply_comment_to_cell(
    ws,
    header_col,
    puic_col,
    header_row,
    puic,
    comment_text,
    data,
    processed,
    skipped,
    not_found,
):
    target_row = find_row_by_value(ws, puic_col, puic, header_row + 1)
    if not target_row:
        not_found.append({**data, "Reason": f"Puic '{puic}' not found"})
        return

    cell = ws.cell(row=target_row, column=header_col)

    style_name = extract_display_text(data.get("Link"))
    if style_name in [s for s in ws.parent.named_styles]:
        cell.style = style_name
    else:
        print(f"Style '{style_name}' not found.")

    is_header_comment = data.get("CommentRow") == header_row
    if is_header_comment:
        header_cell = ws.cell(row=header_row, column=header_col)
        existing_comment = header_cell.comment
        existing_comment_text = (
            existing_comment.text.strip() if existing_comment else ""
        )

        if comment_text == existing_comment_text:
            pass

        if comment_text:
            header_cell.comment = Comment(comment_text, "IMX Tool")

        header_cell.style = style_name

    header_comment = ws.cell(row=header_row, column=header_col).comment
    header_comment_text = header_comment.text.strip() if header_comment else ""
    # Only append once
    if not is_header_comment:
        if comment_text == header_comment_text:
            processed.append(
                {
                    **data,
                    "Reason": "header comment, only cell color",
                    "Value": cell.value,
                }
            )
        elif comment_text:
            cell.comment = Comment(comment_text, "open-imx-comment-replacer")
            processed.append({**data, "Value": cell.value})
        else:
            skipped.append({**data, "Value": cell.value, "Reason": "Empty comment"})
    else:
        # Only append for header comments if non-empty
        if comment_text:
            processed.append({**data, "Value": cell.value})
        else:
            skipped.append({**data, "Value": cell.value, "Reason": "Empty header comment"})



def auto_resize_columns(ws: Worksheet):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add padding
        ws.column_dimensions[col_letter].width = adjusted_width


def apply_comments_from_issue_list(
    issue_list_path: str, new_diff_path: str, output_path: str, header_row: int = 1
):
    issue_wb = load_workbook(issue_list_path, data_only=False)
    diff_wb = load_workbook(new_diff_path)

    if ISSUE_LIST_SHEET_NAME not in issue_wb.sheetnames:
        raise ValueError(f"No '{ISSUE_LIST_SHEET_NAME}' sheet found in the issue list workbook.")

    issue_ws = issue_wb[ISSUE_LIST_SHEET_NAME]
    headers = get_sheet_headers(issue_ws, header_row)

    processed, skipped, not_found = [], [], []

    # Extract all data rows into dictionaries
    all_rows = []
    for row in issue_ws.iter_rows(min_row=header_row + 1, values_only=True):
        data = {
            key: row[idx - 1] if idx - 1 < len(row) else None
            for key, idx in headers.items()
        }
        all_rows.append(data)

    # Sort by CommentRow (as int, if present)
    all_rows.sort(
        key=lambda d: (
            int(d.get("CommentRow"))
            if d.get("CommentRow") is not None
            else float("inf")
        )
    )

    # Process each row after sorting
    for data in all_rows:
        try:
            sheetname = data.get("CommentSheetName")
            imx_path = data.get("ImxPath")
            puic = data.get("Puic")
            comment_text = (data.get("Comment") or "").strip()

            if not all([sheetname, imx_path, puic]):
                not_found.append(
                    {**data, "Reason": "Missing CommentSheetName, ImxPath, or Puic"}
                )
                continue

            if sheetname not in diff_wb.sheetnames:
                not_found.append(
                    {**data, "Reason": f"Sheet '{sheetname}' not found in new diff"}
                )
                continue

            ws = diff_wb[sheetname]
            header_col = find_column_by_value(ws, imx_path, header_row)
            if not header_col:
                not_found.append(
                    {**data, "Reason": f"ImxPath '{imx_path}' not found in header"}
                )
                continue

            puic_col = find_column_by_value(ws, "@puic", header_row)
            if not puic_col:
                not_found.append({**data, "Reason": "@puic column not found"})
                continue

            apply_comment_to_cell(
                ws,
                header_col,
                puic_col,
                header_row,
                puic,
                comment_text,
                data,
                processed,
                skipped,
                not_found,
            )

        except Exception as e:
            fallback = {"Reason": f"Unexpected error: {str(e)}"}
            fallback.update(data)
            not_found.append(fallback)

    summary_ws = create_summary_sheet(diff_wb, processed, skipped, not_found)
    issue_list_ws = diff_wb.create_sheet(ISSUE_LIST_SHEET_NAME)
    copy_full_sheet(issue_ws, issue_list_ws)

    diff_wb._sheets.remove(issue_list_ws)
    diff_wb._sheets.insert(2, issue_list_ws)

    # Reorder sheets
    sheet_order = []
    if "info" in diff_wb.sheetnames:
        sheet_order.append(diff_wb["info"])
    sheet_order.extend([issue_list_ws, summary_ws])
    for sheet in diff_wb.worksheets:
        if sheet not in sheet_order:
            sheet_order.append(sheet)
    diff_wb._sheets = sheet_order


    # Get the 'info' worksheet by name
    info_sheet = diff_wb["info"]

    # Set it as the active sheet using its index
    diff_wb.active = diff_wb.sheetnames.index("info")

    # Mark only this sheet as selected
    for sheet in diff_wb.worksheets:
        sheet.sheet_view.tabSelected = False
    info_sheet.sheet_view.tabSelected = True

    diff_wb.save(output_path)
    print(f"✅ Comments copied and saved to '{output_path}'")
    print(
        f"Summary: {len(processed)} placed, {len(skipped)} skipped, {len(not_found)} failed."
    )
