import shutil
from pathlib import Path

import openpyxl

from imxTools.comments.comments_enums import CommentColumns
from imxTools.comments.comments_extractor import extract_comments_to_new_sheet
from imxTools.comments.comments_replacer import apply_comments_from_issue_list
from imxTools.settings import ISSUE_LIST_SHEET_NAME




def test_comment_sheet_is_present(out_file: str | Path):
    wb = openpyxl.load_workbook(out_file)
    comment_sheet_name = ISSUE_LIST_SHEET_NAME
    assert comment_sheet_name in wb.sheetnames, "Comment sheet not found in workbook"
    comment_sheet = wb[comment_sheet_name]

    expected_headers = CommentColumns.to_dict().keys()
    actual_headers = [cell.value for cell in next(comment_sheet.iter_rows(min_row=1, max_row=1))]
    for header in expected_headers:
        assert header in actual_headers, f"Missing header: {header}"

    assert comment_sheet.max_row > 1, "No comments found in comment sheet"


def test_extract_comments_to_new_file(diff_report, output_path):
    out_file = Path(output_path) / 'comment-list.xlsx'
    extract_comments_to_new_sheet(
        diff_report, f"{out_file}",
        overwrite=True
    )
    assert out_file.exists(), f"Excel file not found at {out_file}"
    test_comment_sheet_is_present(out_file)


def test_extract_comments_to_new_sheet_with_add_to_wb(diff_report, output_path):
    output_file = Path(output_path) / "copied-with-comments.xlsx"

    extract_comments_to_new_sheet(
        diff_report,
        str(output_file),
        add_to_wb=True,
        overwrite=True,
    )
    assert output_file.exists(), f"Excel file not found at {output_file}"
    test_comment_sheet_is_present(output_file)


def test_extract_comments_to_new_sheet_without_overwrite(diff_report, output_path):
    temp_report = Path(output_path) / Path(diff_report).name
    shutil.copy(diff_report, temp_report)

    extract_comments_to_new_sheet(
        temp_report,
        add_to_wb=True,
        overwrite=True  # Explicit that we're overwriting the copy
    )

    assert temp_report.exists(), f"Excel file not found at {temp_report}"
    test_comment_sheet_is_present(temp_report)










def test_apply_comments_to_new_file(diff_report_edit, comment_list, output_path):
    # TODO: comment sheet seems not to be on the second sheet
    # TODO: named styles are not only background fill
    # TODO: seems to add it to the original file
    apply_comments_from_issue_list(
        issue_list_path=comment_list,
        new_diff_path=diff_report_edit,
        output_path=f"{Path(output_path) / 'updated_diff_with_comments.xlsx'}",
    )
